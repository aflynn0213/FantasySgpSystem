"""PointsProcessor — combines PointsHitters + PointsPitchers results,
assigns replacement levels, and produces the final ranked DataFrames.

Mirrors SgpProcessor.  All position-eligibility and replacement-level
logic is identical; the only mechanical differences are:
  • sort/sum key → ``Total_PTS`` (not ``Total_SGP``)
  • col selection → dynamic ``PTS_*`` detection (not hardcoded iloc slices)
  • export column list → PTS_* columns
"""
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

from utils.common_utils import build_config_hitter_counts, get_auction_dollars_spread, get_pitcher_counts, get_position_priority, get_repo_root, is_gcs_enabled, upload_to_bucket, validate_export_columns


class PointsProcessor:
    """Orchestrates end-to-end points-league ranking.

    Args:
        points_hitters:  Initialized PointsHitters instance.
        points_pitchers: Initialized PointsPitchers instance.
        ip_adj:          Optional IP-adjustment system name; appended to the output filename.
    """

    def __init__(self, points_hitters, points_pitchers, ip_adj=None) -> None:
        print("Initializing PointsProcessor...")
        self.weeks = points_hitters.weeks
        temp_hitters = points_hitters.points_df.copy()
        temp_pitchers = points_pitchers.points_df.copy()
        temp_auc_hit = points_hitters.auc_calc.copy()
        ip_adj_part = f"_ip{ip_adj}" if ip_adj else ""
        self.suffix = f"{points_hitters.proj}_{points_pitchers.proj}{ip_adj_part}"

        self.sufficient_pos_counts, self.position_mapping, self.ind_slot_limits, self.comp_slot_limits = build_config_hitter_counts()

        print("Preparing hitter points data...")
        self.hitters_df = self.prepare_data(
            temp_hitters, "Hitter", points_hitters.sb_included, temp_auc_hit
        )
        print("Preparing pitcher points data...")
        self.pitchers_df = self.prepare_data(temp_pitchers, "Pitcher")

        self.hitters_df.sort(by="VAR", ascending=False, inplace=True)
        self.pitchers_df.sort(by="VAR", ascending=False, inplace=True)
        
        self.hitters_df, self.pitchers_df = self.prepare_dollar_values(self.hitters_df, self.pitchers_df)

        cols_in_combined_df = ["Name", "PlayerId", "Total_PTS", "RL", "VAR", "$"]
        sorter = "$"
        if self.weeks < 26:
            cols_in_combined_df = cols_in_combined_df[:3] + cols_in_combined_df[4:]
            sorter = "Total_PTS"
        
        print("Combining data for final points rankings...")
        self.combined_df = (
            pd.concat(
                [
                    self.hitters_df[cols_in_combined_df],
                    self.pitchers_df[cols_in_combined_df],
                ]
            )
            .groupby(["Name", "PlayerId"], as_index=False)
            .sum()
            .sort_values(by=sorter, ascending=False)
        )

        # ADP must be merged back after groupby (summing it would be meaningless)
        adp_source = pd.concat([
            self.hitters_df[["Name", "PlayerId", "ADP"]] if "ADP" in self.hitters_df.columns else pd.DataFrame(columns=["Name", "PlayerId", "ADP"]),
            self.pitchers_df[["Name", "PlayerId", "ADP"]] if "ADP" in self.pitchers_df.columns else pd.DataFrame(columns=["Name", "PlayerId", "ADP"]),
        ]).drop_duplicates("PlayerId")
        if not adp_source.empty and "ADP" in adp_source.columns:
            self.combined_df = self.combined_df.merge(adp_source[["PlayerId", "ADP"]], on="PlayerId", how="left")

        print("Points data prepared!")

    def prepare_dollar_values(self, hitters, pitchers):
        hitter_rostered_sum = hitters[hitters["VAR"] > 0]["VAR"].sum()
        pitcher_rostered_sum = pitchers[pitchers["VAR"] > 0]["VAR"].sum()
        total_rostered_sum = hitter_rostered_sum + pitcher_rostered_sum

        teams, starters, relievers = get_pitcher_counts()
        dollars, hitter_pitcher_split = get_auction_dollars_spread()

        league_dollars = teams * dollars
        hitter_subset = league_dollars * hitter_pitcher_split
        pitcher_subset = league_dollars - hitter_subset
        hitter_dollar_by_points = float(hitter_subset) / hitter_rostered_sum
        pitcher_dollar_by_points = float(pitcher_subset) / pitcher_rostered_sum

        hitters["$"] = hitters["VAR"] * hitter_dollar_by_points
        pitchers["$"] = pitchers["VAR"] * pitcher_dollar_by_points
        
        return hitters,pitchers

    # ------------------------------------------------------------------
    # Data preparation
    # ------------------------------------------------------------------

    def prepare_data(self, df, player_type, sb_included=False, auc_calc=None):
        df = df.reset_index()

        # Identify PTS_* score columns dynamically
        pts_cols = [c for c in df.columns if c.startswith("PTS_")]

        if player_type == "Hitter":
            # Optionally exclude SB from total
            active_pts = pts_cols if sb_included else [c for c in pts_cols if c != "PTS_SB"]

            df["Total_PTS"] = df[active_pts].sum(axis=1)
            df["Total_PTS_wSB"] = df[pts_cols].sum(axis=1)

            df = df.sort_values(by="Total_PTS", ascending=False).reset_index(drop=True)

            if self.weeks == 26:
                auc_calc = auc_calc.rename(columns={"POS": "ELIG"})
                df = df.merge(auc_calc[["PlayerId", "ELIG"]], on="PlayerId", how="left")

                temp_path = os.path.join(get_repo_root(), 'temp_players_pts.xlsx')
                df.to_excel(temp_path)
                if is_gcs_enabled():
                    upload_to_bucket(temp_path, 'debug/temp_players_pts.xlsx')

                print("Computing replacement level (RL) for points hitters...")
                df = self.compute_replacement_level(df)

        else:  # Pitcher
            df["Total_PTS"] = df[pts_cols].sum(axis=1)
            df["Starter"] = np.where(df["GS"] > 5, 1, 0)
            df["POS"] = np.where(df["GS"] > 5, "SP", "RP")

            df = df.sort_values(by="Total_PTS", ascending=False).reset_index(drop=True)

            print("Computing replacement level for points pitchers...")
            num_teams, num_sp, num_rp = get_pitcher_counts()
            starter_rl = df[df["Starter"] == 1]["Total_PTS"].iloc[num_teams * num_sp]
            reliever_rl = df[df["Starter"] == 0]["Total_PTS"].iloc[num_teams * num_rp]

            df["RL"] = df.apply(
                lambda row: starter_rl if row["Starter"] == 1 else reliever_rl, axis=1
            )
            df["VAR"] = df["Total_PTS"] - df["RL"]

        print(f"[✔] {player_type} points data preparation complete!")
        return df

    def compute_replacement_level(self, df):
        """Compute RL per position using the rostered universe — mirrors SgpProcessor."""
        print("Identifying rostered universe (points)...")

        def get_rostered_universe(df):
            INDIVIDUAL  = list(self.ind_slot_limits.keys())
            COMPOSITES  = list(self.comp_slot_limits.keys())
            IND_TO_COMP = {ind: grp for ind, grp in self.position_mapping.items() if grp in self.comp_slot_limits}

            ind_limits    = self.ind_slot_limits
            comp_limits   = self.comp_slot_limits
            util_limit    = self.sufficient_pos_counts.get('UTIL', 0)
            priority_rank = {p: i for i, p in enumerate(get_position_priority())}

            ind_counts  = {p: 0 for p in INDIVIDUAL}
            comp_counts = {p: 0 for p in COMPOSITES}
            util_count  = 0
            rostered    = []

            df_sorted = df.sort_values(by="Total_PTS", ascending=False).reset_index(
                drop=True
            )

            for _, row in df_sorted.iterrows():
                raw  = [p.strip() for p in str(row["ELIG"]).split("/") if p.strip() in ind_limits]
                elig = sorted(raw, key=lambda p: priority_rank.get(p, 99))
                assigned = False

                # For each elig position in priority order:
                #   1. try its individual slot
                #   2. if that's full, try its composite flex slot
                #   then move to the next elig position
                for pos in elig:
                    if ind_counts[pos] < ind_limits[pos]:
                        ind_counts[pos] += 1
                        r = row.copy()
                        r["ASSIGNED_POS"] = pos
                        rostered.append(r)
                        assigned = True
                        break
                    comp = IND_TO_COMP.get(pos)
                    if comp and comp_counts[comp] < comp_limits[comp]:
                        comp_counts[comp] += 1
                        r = row.copy()
                        r["ASSIGNED_POS"] = f"{pos}({comp} flex)"
                        rostered.append(r)
                        assigned = True
                        break

                if not assigned and util_count < util_limit:
                    util_count += 1
                    r = row.copy()
                    r["ASSIGNED_POS"] = "UTIL"
                    rostered.append(r)

                ind_full  = all(ind_counts[p] >= ind_limits[p] for p in INDIVIDUAL if ind_limits[p] > 0)
                comp_full = all(comp_counts[p] >= comp_limits[p] for p in COMPOSITES)
                if ind_full and comp_full and util_count >= util_limit:
                    break

            return (
                pd.DataFrame(rostered)
                .sort_values(by="Total_PTS", ascending=False)
                .reset_index(drop=True)
            )

        rostered_df = get_rostered_universe(df)

        # Debug workbook: ALL players, rostered rows highlighted green, ASSIGNED_POS visible
        debug_path   = os.path.join(get_repo_root(), "rostered_pts.xlsx")
        rostered_ids = set(rostered_df["PlayerId"].astype(str))
        debug_df = df.merge(
            rostered_df[["PlayerId", "ASSIGNED_POS"]].drop_duplicates("PlayerId"),
            on="PlayerId", how="left"
        )
        with pd.ExcelWriter(debug_path, engine="openpyxl") as writer:
            debug_df.to_excel(writer, sheet_name="Players", index=False)
            ws = writer.sheets["Players"]
            from openpyxl.styles import PatternFill
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            pid_col = [c.value for c in ws[1]].index("PlayerId") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if str(row[pid_col - 1].value) in rostered_ids:
                    for cell in row:
                        cell.fill = green_fill

            # Bucket Counts sheet: cumulative fill per bucket in assignment order
            fill_df = (
                pd.get_dummies(rostered_df["ASSIGNED_POS"])
                .cumsum()
            )
            fill_df.insert(0, "Name", rostered_df["Name"].values)
            fill_df.insert(1, "ASSIGNED_POS", rostered_df["ASSIGNED_POS"].values)
            fill_df.insert(2, "ELIG", rostered_df["ELIG"].values)
            fill_df.to_excel(writer, sheet_name="Bucket Counts", index=False)

        if is_gcs_enabled():
            upload_to_bucket(debug_path, 'debug/rostered_pts.xlsx')

        print(f"Rostered universe identified ({len(rostered_df)} players).")

        print("Finding worst rostered / best non-rostered per position...")
        worst_rostered = {}
        for pos in ["1B", "3B", "2B", "SS", "C", "OF"]:
            at_pos = rostered_df[
                rostered_df["ASSIGNED_POS"].str.startswith(pos, na=False) |
                (
                    (rostered_df["ASSIGNED_POS"] == "UTIL") &
                    rostered_df["ELIG"].str.contains(pos, na=False)
                )
            ]
            if at_pos.empty:
                print(f"[WARN] No rostered players found for {pos}")
                continue
            worst_player = at_pos.nsmallest(1, "Total_PTS")
            worst_rostered[pos] = worst_player["Total_PTS"].values[0]
            print(f"[DEBUG] Worst rostered {pos}: {worst_player['Name'].values[0]} (PTS: {worst_rostered[pos]:.1f})")

        best_replacements = {}
        for pos in worst_rostered:
            best_player = df[
                ~df["PlayerId"].isin(rostered_df["PlayerId"])
                & df["ELIG"].str.contains(pos, na=False)
            ].nlargest(1, "Total_PTS")
            if not best_player.empty:
                best_replacements[pos] = best_player["Total_PTS"].values[0]
                print(
                    f"[DEBUG] Best non-rostered {pos}: {best_player['Name'].values[0]} "
                    f"(PTS: {best_replacements[pos]:.1f})"
                )
            else:
                best_replacements[pos] = float("-inf")

        rl_values = {
            pos: (best_replacements[pos] + worst_rostered[pos]) / 2
            if worst_rostered[pos] != float("inf")
            and best_replacements[pos] != float("-inf")
            else float("inf")
            for pos in worst_rostered
        }

        max_worst = max(
            (v for v in worst_rostered.values() if v != float("inf")), default=float("inf")
        )
        max_best = max(
            (v for v in best_replacements.values() if v != float("-inf")), default=float("-inf")
        )
        rl_values["DH"] = (max_best + max_worst) / 2
        print(f"[DEBUG] DH RL penalized to: {rl_values['DH']} (Max of all worst and best non-rostered)")

        df["RL"] = df["ELIG"].apply(
            lambda elig: min(
                rl_values.get(pos, float("inf")) for pos in elig.split("/")
            )
        )
        df["VAR"] = df["Total_PTS"] - df["RL"]

        print("Points replacement level calculation complete!")
        return df

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_points(self, sb: bool) -> str:
        """Write a combined Excel workbook with Hitters / Pitchers / Combined sheets."""
        print("Exporting Points Results...")

        SAVE_FOLDER = os.path.join(get_repo_root(), "results")
        os.makedirs(SAVE_FOLDER, exist_ok=True)

        sb_string = "_sb_included" if sb else ""
        base_name = f"Points_Results_{self.suffix}{sb_string}.xlsx"
        file_name = os.path.join(SAVE_FOLDER, base_name)

        # Dynamic PTS_ columns for each tab
        hit_pts_cols = [c for c in self.hitters_df.columns if c.startswith("PTS_")]
        pit_pts_cols = [c for c in self.pitchers_df.columns if c.startswith("PTS_")]

        hit_export = ["Name", "PlayerId", "PA"] + hit_pts_cols + ["Total_PTS_wSB", "Total_PTS", "RL", "VAR"]
        pit_export = ["Name", "PlayerId", "POS", "GS", "IP"] + pit_pts_cols + ["Total_PTS", "RL", "VAR"]

        # Insert optional columns right after PlayerId when present
        if "ADP" in self.hitters_df.columns:
            hit_export.insert(2, "ADP")
        if "ADP" in self.pitchers_df.columns:
            pit_export.insert(2, "ADP")
        if "ELIG" in self.hitters_df.columns:
            hit_export.insert(hit_export.index("PA"), "ELIG")

        # Drop cols that don't exist yet (in-season shortcut)
        hit_export = [c for c in hit_export if c in self.hitters_df.columns]
        pit_export = [c for c in pit_export if c in self.pitchers_df.columns]

        validate_export_columns(self.hitters_df, hit_export, "Hitters Points")
        validate_export_columns(self.pitchers_df, pit_export, "Pitchers Points")

        with pd.ExcelWriter(file_name) as writer:
            self.hitters_df[hit_export].to_excel(writer, sheet_name="Hitters", index=False)
            self.pitchers_df[pit_export].to_excel(writer, sheet_name="Pitchers", index=False)
            self.combined_df.to_excel(writer, sheet_name="Combined", index=False)

        # Apply blue/bold formatting to pitchers in the Combined sheet
        wb = load_workbook(file_name)
        ws_pitchers = wb["Pitchers"]
        ws_combined = wb["Combined"]
        blue_underlined = Font(color="0000FF", underline="single", bold=True)

        pitcher_ids = set()
        for row in ws_pitchers.iter_rows(min_row=2, values_only=True):
            name, player_id, *_ = row
            pitcher_ids.add((name, player_id))

        for row in ws_combined.iter_rows(min_row=2, max_row=ws_combined.max_row):
            if (row[0].value, row[1].value) in pitcher_ids:
                for cell in row:
                    cell.font = blue_underlined

        wb.save(file_name)

        if is_gcs_enabled():
            gcs_blob_path = f"results/{base_name}"
            upload_to_bucket(file_name, gcs_blob_path)
            print(f"[GCS] Uploaded {base_name}")

        print(f"Exported Points Results to {file_name}")
        return base_name
