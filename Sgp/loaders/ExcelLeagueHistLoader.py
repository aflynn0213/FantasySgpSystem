from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from openpyxl import load_workbook
from .ILeagueDataLoader import ILeagueDataLoader
    
class ExcelLeagueHistLoader(ILeagueDataLoader):
    """
    Reads a sheet (default 'Parameters') where each row is:
      Category  Replacement  Std
    Contains a hash (data) mapping category -> (replacement_value, std_value).
    """

    def __init__(self, workbook_path: str = "included/leaguehistory.xlsx", sheet_name: str = "Parameters"):
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name

    def load(self) -> Dict[str, Tuple[Optional[float], Optional[float]]]:
        p = str(self.workbook_path)
        wb = load_workbook(Path(p), data_only=True)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {self.sheet_name} not found in workbook {p}")
        sheet = wb[self.sheet_name]

        out: Dict[str, Tuple[Optional[float], Optional[float]]] = {}

        # data begins at the first row (no header) or optionally a header row
        # read first three columns A,B,C
        for row in sheet.iter_rows(min_row=1, max_col=3):
            key_cell = row[0].value  # column A
            # skip rows with empty key;
            if key_cell is None:
                continue
            key = str(key_cell).strip()
            if key == "":
                continue

            repl_cell = row[1].value if len(row) > 1 else None
            std_cell = row[2].value if len(row) > 2 else None

            def to_float_safe(x):
                if x is None:
                    return None
                if isinstance(x, (int, float)):
                    return float(x)
                s = str(x).strip()
                if s == "":
                    return None
                try:
                    return float(s)
                except ValueError:
                    return None

            repl_val = to_float_safe(repl_cell)
            std_val = to_float_safe(std_cell)

            out[key] = (repl_val, std_val)

        self.data = out

