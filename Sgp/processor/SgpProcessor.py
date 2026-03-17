import pandas as pd
import numpy as np
import os

from openpyxl import load_workbook
from openpyxl.styles import Font

from utils.common_utils import build_config_hitter_counts, get_position_priority, get_repo_root, is_gcs_enabled, upload_to_bucket, validate_export_columns

class SgpProcessor:
    def __init__(self, sgp_hitters, sgp_pitchers, ip_adj=None):

        print("Initializing SGP Processor...")
        self.weeks = sgp_hitters.weeks
        temp_hitters = sgp_hitters.sgp_df.copy()
        temp_pitchers = sgp_pitchers.sgp_df.copy()
        temp_auc_hit = sgp_hitters.auc_calc.copy()
        ip_adj_part = f"_ip{ip_adj}" if ip_adj else ""
        self.suffix = f"{sgp_hitters.proj}_{sgp_pitchers.proj}{ip_adj_part}"
        print(temp_pitchers)
        self.sufficient_pos_counts, self.position_mapping, self.ind_slot_limits, self.comp_slot_limits = build_config_hitter_counts()
        
         
        print("Preparing hitter data...")
        self.hitters_df = self.prepare_data(temp_hitters, 'Hitter', sgp_hitters.sb_included, temp_auc_hit)
        print("Preparing pitcher data...")
        self.pitchers_df = self.prepare_data(temp_pitchers, 'Pitcher')
        
        cols_in_hitters_df = ['Name', 'PlayerId', 'Total_SGP', 'RL', 'VAR']
        sorter = 'VAR'
        if self.weeks < 26:
            cols_in_hitters_df = cols_in_hitters_df[:3]
            sorter = 'Total_SGP'
        print("Combining data for final rankings...")
        self.combined_df = pd.concat([
            self.hitters_df[cols_in_hitters_df],
            self.pitchers_df[['Name', 'PlayerId', 'Total_SGP', 'RL', 'VAR']]
        ]).groupby(['Name', 'PlayerId'], as_index=False).sum().sort_values(by=sorter, ascending=False)

        # ADP must be merged back after groupby (summing it would be meaningless)
        adp_source = pd.concat([
            self.hitters_df[['Name', 'PlayerId', 'ADP']] if 'ADP' in self.hitters_df.columns else pd.DataFrame(columns=['Name', 'PlayerId', 'ADP']),
            self.pitchers_df[['Name', 'PlayerId', 'ADP']] if 'ADP' in self.pitchers_df.columns else pd.DataFrame(columns=['Name', 'PlayerId', 'ADP']),
        ]).drop_duplicates('PlayerId')
        if not adp_source.empty and 'ADP' in adp_source.columns:
            self.combined_df = self.combined_df.merge(adp_source[['PlayerId', 'ADP']], on='PlayerId', how='left')

        print("SGP Data Prepared!")
    
    def prepare_data(self, df, player_type, sb_included=False, auc_calc = None):
        df = df.reset_index()
        if (player_type=='Hitter'):
            cols_included = list(range(2,8))
            if (not sb_included):
                cols_included.remove(5)
            
            df['Total_SGP'] = df.iloc[:, cols_included].sum(axis=1)
            df['Total_SGP_wSB'] = df.iloc[:, 2:8].sum(axis=1)
            
            df = df.sort_values(by="Total_SGP", ascending=False).reset_index()
            
            if self.weeks == 26:
                auc_calc = auc_calc.rename(columns={'POS': 'ELIG'})
                df = df.merge(auc_calc[['PlayerId', 'ELIG']], on='PlayerId', how='left')

                temp_path = os.path.join(get_repo_root(), 'temp_players.xlsx')
                df.to_excel(temp_path)
                if is_gcs_enabled():
                    upload_to_bucket(temp_path, 'debug/temp_players.xlsx')

                print("Computing replacement level (RL)...")
                df = self.compute_replacement_level(df)
            
        else:
            df['Total_SGP'] = df.iloc[:, 2:8].sum(axis=1)
            df['Starter'] = np.where(df['GS'] > 5, 1, 0)
            df['POS'] = np.where(df['GS'] > 5, 'SP', 'RP')
            
            df = df.sort_values(by="Total_SGP", ascending=False).reset_index()
            
            print("Computing replacement level for pitchers...")
            starter_rl = df[df["Starter"] == 1]["Total_SGP"].iloc[96]
            reliever_rl = df[df["Starter"] == 0]["Total_SGP"].iloc[48]
            
            df["RL"] = df.apply(lambda row: starter_rl if row["Starter"] == 1 else reliever_rl, axis=1)
            df["VAR"] = df["Total_SGP"] - df["RL"]
        
        print(f"[✔] {player_type} data preparation complete!")
        return df

    def compute_replacement_level(self, df):
        """Compute RL per position, ensuring required number of rostered players"""
        print("Identifying rostered universe...")
        
        def get_rostered_universe(df):
            """Find the highest-ranked players that fill required positions"""
            INDIVIDUAL  = list(self.ind_slot_limits.keys())
            COMPOSITES  = list(self.comp_slot_limits.keys())
            IND_TO_COMP = {ind: grp for ind, grp in self.position_mapping.items() if grp in self.comp_slot_limits}

            ind_limits    = self.ind_slot_limits
            comp_limits   = self.comp_slot_limits
            util_limit    = self.sufficient_pos_counts.get('UTIL', 0)
            priority_rank = {p: i for i, p in enumerate(get_position_priority())}

            ind_counts  = {p: 0 for p in INDIVIDUAL}
            comp_counts = {p: 0 for p in COMPOSITES}
            util_count  = 0
            rostered    = []

            df_sorted = df.sort_values(by="Total_SGP", ascending=False).reset_index(drop=True)

            for _, row in df_sorted.iterrows():
                # All positions this player is eligible for, sorted by priority
                raw  = [p.strip() for p in str(row["ELIG"]).split("/") if p.strip() in ind_limits]
                elig = sorted(raw, key=lambda p: priority_rank.get(p, 99))
                assigned = False

                # For each elig position in priority order:
                #   1. try its individual slot
                #   2. if that's full, try its composite flex slot
                #   then move to the next elig position
                for pos in elig:
                    if ind_counts[pos] < ind_limits[pos]:
                        ind_counts[pos] += 1
                        r = row.copy()
                        r["ASSIGNED_POS"] = pos
                        rostered.append(r)
                        assigned = True
                        break
                    comp = IND_TO_COMP.get(pos)
                    if comp and comp_counts[comp] < comp_limits[comp]:
                        comp_counts[comp] += 1
                        r = row.copy()
                        r["ASSIGNED_POS"] = f"{pos}({comp} flex)"
                        rostered.append(r)
                        assigned = True
                        break

                # All elig slots exhausted — try UTIL
                if not assigned and util_count < util_limit:
                    util_count += 1
                    r = row.copy()
                    r["ASSIGNED_POS"] = "UTIL"
                    rostered.append(r)

                # Stop only when every bucket is genuinely full
                ind_full  = all(ind_counts[p] >= ind_limits[p] for p in INDIVIDUAL if ind_limits[p] > 0)
                comp_full = all(comp_counts[p] >= comp_limits[p] for p in COMPOSITES)
                if ind_full and comp_full and util_count >= util_limit:
                    break

            # Convert to DataFrame and sort by Total_SGP before returning
            return pd.DataFrame(rostered).sort_values(by="Total_SGP", ascending=False).reset_index(drop=True)

        rostered_df = get_rostered_universe(df)

        # Debug workbook: ALL players, rostered rows highlighted green
        debug_path = os.path.join(get_repo_root(), "rostered.xlsx")
        rostered_ids = set(rostered_df["PlayerId"].astype(str))
        debug_df = df.merge(
            rostered_df[["PlayerId", "ASSIGNED_POS"]].drop_duplicates("PlayerId"),
            on="PlayerId", how="left"
        )
        with pd.ExcelWriter(debug_path, engine="openpyxl") as writer:
            # Players sheet: all players with ASSIGNED_POS, rostered highlighted green
            debug_df.to_excel(writer, sheet_name="Players", index=False)
            ws = writer.sheets["Players"]
            from openpyxl.styles import PatternFill
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            pid_col = [c.value for c in ws[1]].index("PlayerId") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if str(row[pid_col - 1].value) in rostered_ids:
                    for cell in row:
                        cell.fill = green_fill

            # Bucket Counts sheet: cumulative fill per bucket in assignment order
            fill_df = (
                pd.get_dummies(rostered_df["ASSIGNED_POS"])
                .cumsum()
            )
            fill_df.insert(0, "Name", rostered_df["Name"].values)
            fill_df.insert(1, "ASSIGNED_POS", rostered_df["ASSIGNED_POS"].values)
            fill_df.insert(2, "ELIG", rostered_df["ELIG"].values)
            fill_df.to_excel(writer, sheet_name="Bucket Counts", index=False)

        if is_gcs_enabled():
            upload_to_bucket(debug_path, 'debug/rostered.xlsx')

        print(f"Rostered universe identified ({len(rostered_df)} players).")

        print("Finding worst rostered player per position...")
        worst_rostered = {}

        for pos in ["1B", "3B", "2B", "SS", "C", "OF"]:
            # A player counts as rostered at pos if:
            #   - directly assigned to the individual slot (ASSIGNED_POS == pos)
            #   - assigned to that position's composite flex (ASSIGNED_POS starts with "pos(")
            #   - assigned to UTIL and eligible at pos (they are the marginal player at that pos)
            at_pos = rostered_df[
                rostered_df["ASSIGNED_POS"].str.startswith(pos, na=False) |
                (
                    (rostered_df["ASSIGNED_POS"] == "UTIL") &
                    rostered_df["ELIG"].str.contains(pos, na=False)
                )
            ]
            if at_pos.empty:
                print(f"[WARN] No rostered players found for {pos}")
                continue
            worst_player = at_pos.nsmallest(1, "Total_SGP")
            worst_rostered[pos] = worst_player["Total_SGP"].values[0]
            print(f"[DEBUG] Worst rostered {pos}: {worst_player['Name'].values[0]} (SGP: {worst_rostered[pos]})")
            
        print("Finding best non-rostered player per position...")
        best_replacements = {}
        for pos in worst_rostered.keys():
            best_player = df[~df["PlayerId"].isin(rostered_df["PlayerId"]) & df["ELIG"].str.contains(pos, na=False)].nlargest(1, "Total_SGP")
            if not best_player.empty:
                best_replacements[pos] = best_player["Total_SGP"].values[0]
                print(f"[DEBUG] Best non-rostered {pos}: {best_player['Name'].values[0]} (SGP: {best_replacements[pos]})")
            else:
                best_replacements[pos] = float("-inf")

        print("Computing RL for each position...")
        rl_values = {
            pos: (best_replacements[pos] + worst_rostered[pos]) / 2
            if worst_rostered[pos] != float("inf") and best_replacements[pos] != float("-inf")
            else float("inf")
            for pos in worst_rostered
        }

        max_worst = max([val for val in worst_rostered.values() if val != float("inf")], default=float("inf"))
        max_best = max([val for val in best_replacements.values() if val != float("-inf")], default=float("-inf"))
    
        rl_values["DH"] = (max_best+max_worst)/2  # Take the maximum of both

        print(f"[DEBUG] DH RL penalized to: {rl_values['DH']} (Max of all worst and best non-rostered)")
        
        df["RL"] = df["ELIG"].apply(lambda elig: min(rl_values.get(pos, float("inf")) for pos in elig.split("/")))
        df["VAR"] = df["Total_SGP"] - df["RL"]
        
        print("Replacement level calculation complete!")
        return df
    
    def export_sgp(self, sb):
        print("Exporting SGP Results...")

        SAVE_FOLDER = os.path.join(get_repo_root(), "results")
        os.makedirs(SAVE_FOLDER, exist_ok=True)

        sb_string = "_sb_included" if sb else ""
        base_name = f"SGP_Results_{self.suffix}{sb_string}.xlsx"
        file_name = os.path.join(SAVE_FOLDER, base_name)

        hit_cols = ['Name', 'PlayerId', 'PA', 'SGP_R', 'SGP_HR', 'SGP_RBI', 'SGP_SB', 'SGP_OBP', 'SGP_SLG', 'Total_SGP_wSB', 'Total_SGP', 'RL', 'VAR']
        pit_cols = ['Name', 'PlayerId', 'POS', 'GS', 'IP', 'SGP_SO', 'SGP_QS', 'SGP_SV_HLD', 'SGP_ERA', 'SGP_WHIP', 'SGP_K/BB', 'Total_SGP', 'RL', 'VAR']

        # Insert optional columns right after PlayerId when present
        if 'ADP' in self.hitters_df.columns:
            hit_cols.insert(2, 'ADP')
        if 'ADP' in self.pitchers_df.columns:
            pit_cols.insert(2, 'ADP')
        if 'ELIG' in self.hitters_df.columns:
            hit_cols.insert(hit_cols.index('PA'), 'ELIG')

        validate_export_columns(self.hitters_df, hit_cols, "Hitters SGP")
        validate_export_columns(self.pitchers_df, pit_cols, "Pitchers SGP")

        with pd.ExcelWriter(file_name) as writer:
            self.hitters_df[hit_cols].to_excel(writer, sheet_name='Hitters', index=False)
            self.pitchers_df[pit_cols].to_excel(writer, sheet_name='Pitchers', index=False)
            self.combined_df.to_excel(writer, sheet_name='Combined', index=False)

        wb = load_workbook(file_name)
        ws_pitchers = wb["Pitchers"]  # Select the "Pitchers" sheet
        ws_combined = wb["Combined"]  # Select the "Combined" sheet

        # Define formatting (blue font & underlined)
        blue_underlined = Font(color="0000FF", underline="single", bold=True)

        # Get the list of pitchers from the "Combined" sheet (Name & PlayerId)
        combined_pitchers = set()
        for row in ws_pitchers.iter_rows(min_row=2, values_only=True):  # Skip header
            name, player_id, *_ = row  # Extract Name & PlayerId
            combined_pitchers.add((name, player_id))

        # Apply formatting only to pitchers who appear in the Combined sheet
        for row in ws_combined.iter_rows(min_row=2, max_row=ws_combined.max_row):
            name_cell, player_id_cell = row[0], row[1]  # Assuming Name is in column A, PlayerId in B
            if (name_cell.value, player_id_cell.value) in combined_pitchers:
                for cell in row:  # Apply formatting to the whole row
                    cell.font = blue_underlined

        wb.save(file_name)

        if is_gcs_enabled():
            gcs_blob_path = f"results/{base_name}"
            upload_to_bucket(file_name, gcs_blob_path)
            print(f"[GCS] Uploaded {base_name}")

        print(f"Exported SGP Results to {file_name}")
        return base_name
