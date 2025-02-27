from asyncio.windows_events import NULL
import string
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from processor.SgpProcessor import SgpProcessor 

NUM_TEAMS = 12
NUM_BATS = 13
NUM_STARTERS = 9
NUM_RELIEVERS = 3
class SgpHitters():
    def __init__(self,proj) -> None:
        wb = load_workbook('LeagueStatsSGPInvest.xlsm', data_only=True)
        self.proj = proj.split()[0]
        
        self.stats = pd.read_excel(f'projections/fangraphs_hitting_{self.proj}.xlsx', sheet_name=0)    
        
        sheet = wb["3YR RUNNING AVG SGP"]

        self.replacement_levels = {
            'R': sheet['Q26'].value, 'HR': sheet['R26'].value, 'RBI': sheet['S26'].value, 'SB': sheet['T26'].value,
            'OBP': sheet['U26'].value, 'SLG': sheet['V26'].value
        }
    
        self.cat_stds = {
            'R': sheet['Q27'].value, 'HR': sheet['R27'].value, 'RBI': sheet['S27'].value, 'SB': sheet['T27'].value,
            'OBP': sheet['U27'].value, 'SLG': sheet['V27'].value
        }
        
        self.team_opportunities = {}
        self.team_value = {}
        
        self.team_rate_values_processing()
        
        self.process_hitters_sgp()
        
        self.sgp_df['PA'] = self.stats['PA'] - self.stats['SH']
        
        self.sgp_df[['Name', 'PlayerId']] = self.stats[['Name', 'PlayerId']]
        self.sgp_df.set_index(['Name','PlayerId'], inplace=True)
        
        
    def cat_calc_sgp(self,projection:string):
        return (self.stats[projection] - self.replacement_levels[projection]) / self.cat_stds[projection]

    def rate_calc_sgp(self,cat:string,opportunities:string):
        if opportunities == 'AB':
            team_cat = 'SLG_AB'
            player_opps = self.stats[opportunities]
        elif opportunities == 'PA':
            team_cat = 'OBP_PA'
            player_opps = self.stats[opportunities] - self.stats['SH']

        player_val = self.stats[cat]*player_opps
        team_val_wo_average_player = self.team_value[team_cat]
        total_opps = self.team_opportunities[opportunities] + player_opps

        return ((team_val_wo_average_player+player_val)/(total_opps) - self.replacement_levels[cat])/self.cat_stds[cat]

    def process_hitters_sgp(self):
        self.sgp_df = pd.DataFrame()
        for cat in ['R', 'HR', 'RBI', 'SB']:
            self.sgp_df[f'SGP_{cat}'] = self.cat_calc_sgp(cat)
        
        for cat, opps in [('OBP', 'PA'), ('SLG', 'AB')]:
            self.sgp_df[f'SGP_{cat}'] = self.rate_calc_sgp(cat,opps)
        
    def team_rate_values_processing(self):
        temp_df = pd.read_excel(f"auction_calculator_exports/auc_calc_hitting_{self.proj}.xlsx",sheet_name=0) 
        temp_df = temp_df.merge(    self.stats[['PlayerId', 'SH', 'AB']],  
                                    on='PlayerId', 
                                    how='left'
                                )
        print(temp_df)
        
        for cat,val,cat_val in [('OBP','PA','OBP_PA'), ('SLG','AB','SLG_AB')]: 
            if val == 'PA':
                temp_df['PA_SH'] = temp_df['PA'] - temp_df['SH']
                avg_opps = temp_df['PA_SH'].head(NUM_BATS*NUM_TEAMS).mean()
            elif val == 'AB':
                avg_opps = temp_df['AB'].head(NUM_BATS*NUM_TEAMS).mean()
                
            avg_team_opps_wo_replacement = avg_opps*(NUM_BATS-1)
            avg_team_value_wo_replacement = avg_team_opps_wo_replacement*self.replacement_levels[cat]

            self.team_opportunities[val] = avg_team_opps_wo_replacement
            self.team_value[cat_val] = avg_team_value_wo_replacement
        
class SgpPitchers():
    def __init__(self,proj:string):
        wb = load_workbook('LeagueStatsSGPInvest.xlsm', data_only=True)
        self.proj = proj.split()[0].lower()
        
        self.stats = pd.read_excel(f'projections/fangraphs_pitching_{self.proj}.xlsx', sheet_name=0)
        
        sheet = wb["3YR RUNNING AVG SGP"]

        self.replacement_levels = {
            'SO': sheet['W26'].value,'QS': sheet['X26'].value, 'SV_HLD': sheet['AB26'].value, 
            'ERA': sheet['Y26'].value,'WHIP': sheet['Z26'].value, 'K/BB': sheet['AA26'].value
        }
    
        self.cat_stds = {
            'SO': sheet['W27'].value,'QS': sheet['X27'].value, 'SV_HLD': sheet['AB27'].value, 
            'ERA': sheet['Y27'].value, 'WHIP': sheet['Z27'].value, 'K/BB': sheet['AA27'].value
        }
        
        
        self.team_opportunities = {}
        self.team_value = {}
        
        self.team_rate_values_processing()
        
        self.process_pitchers_sgp()
        
        self.sgp_df['IP'] = self.stats['IP']
        self.sgp_df['GS'] = self.stats['GS']
        
        self.sgp_df[['Name', 'PlayerId']] = self.stats[['Name', 'PlayerId']]
        self.sgp_df.set_index(['Name','PlayerId'], inplace=True)
        
    def cat_calc_sgp(self,projection,cat:string):
        return (projection - self.replacement_levels[cat]) / self.cat_stds[cat]

    def pitcher_rate_calc(self,projection,cat,opps):
        multiplier = 1
        if (cat == 'ERA'):
            multiplier = 9
            
        team_val_wo_average_player = multiplier*self.team_value[cat]
        total_opps = self.team_opportunities[opps] + self.stats[opps]
        return ((team_val_wo_average_player+projection)/(total_opps) - self.replacement_levels[cat])/self.cat_stds[cat]

    def rate_calc_sgp(self,cat,opps):
        if(cat=='ERA'):
            val = 9*self.stats['ER']
        elif(cat=='WHIP'):
            val = self.stats['H']+self.stats['BB']
        elif(cat=="K/BB"):
            val = self.stats['SO']
        else:
            return NULL
        return self.pitcher_rate_calc(val,cat,opps)

    def process_pitchers_sgp(self):
        self.sgp_df = pd.DataFrame()
        for cat in ['SO', 'QS', 'SV_HLD']:
            if cat == 'SV_HLD':
                val = self.stats['SV'] + self.stats['HLD']
            else:
                val = self.stats[cat]
            self.sgp_df[f'SGP_{cat}'] = self.cat_calc_sgp(val,cat)
        
        for cat, opps in [('ERA','IP'), ('WHIP', 'IP'), ('K/BB', 'BB')]:
            self.sgp_df[f'SGP_{cat}'] = self.rate_calc_sgp(cat,opps)
 
    def team_rate_values_processing(self):
        temp_df = pd.read_excel(f"auction_calculator_exports/auc_calc_pitching_{self.proj}.xlsx",sheet_name=0) 
        
        print(temp_df)
        multiplier = 1
        
        for cat,val in [('ERA','IP'), ('WHIP','IP'), ('K/BB', 'BB')]: 
            if val == 'IP':
                avg_opps = temp_df['IP'].head((NUM_STARTERS+NUM_RELIEVERS)*NUM_TEAMS).mean()
                multiplier = 9 if cat == 'ERA' else 1       
            elif val == 'BB':
                avg_opps = self.replacement_levels['SO']/self.replacement_levels['K/BB']
                
            avg_team_opps_wo_replacement = avg_opps*(NUM_STARTERS+NUM_RELIEVERS-1)
            avg_team_value_wo_replacement = avg_team_opps_wo_replacement*self.replacement_levels[cat]/multiplier

            self.team_opportunities[val] = avg_team_opps_wo_replacement
            self.team_value[cat] = avg_team_value_wo_replacement

if __name__ == "__main__":
    sgp_hit = SgpHitters(proj="atc")
    sgp_pit = SgpPitchers(proj="atc")
    sgp_pit_oopsy = SgpPitchers(proj="oopsy")
    
    processor_atc = SgpProcessor(sgp_hit,sgp_pit)   
    processor_oopsy = SgpProcessor(sgp_hit,sgp_pit_oopsy)
    
    # Write to excel results files for sorted rankings 
    processor_atc.export_sgp()
    processor_oopsy.export_sgp() 
    

