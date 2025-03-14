import pandas as pd
import numpy as np
import os

from openpyxl import load_workbook
from openpyxl.styles import Font

class SgpProcessor:
    def __init__(self, sgp_hitters, sgp_pitchers):

        print("[*] Initializing SGP Processor...")
        
        temp_hitters = sgp_hitters.sgp_df.copy()
        temp_pitchers = sgp_pitchers.sgp_df.copy()
        temp_auc_hit = sgp_hitters.auc_calc.copy()
        self.suffix = f"{sgp_hitters.proj}_{sgp_pitchers.proj}"

        print("[*] Preparing hitter data...")
        self.hitters_df = self.prepare_data(temp_hitters, 'Hitter', sgp_hitters.sb_included, temp_auc_hit)
        print("[*] Preparing pitcher data...")
        self.pitchers_df = self.prepare_data(temp_pitchers, 'Pitcher')
                                                     
        print("[*] Combining data for final rankings...")
        self.combined_df = pd.concat([
            self.hitters_df[['Name', 'PlayerId', 'Total_SGP', 'RL', 'VAR']],
            self.pitchers_df[['Name', 'PlayerId', 'Total_SGP', 'RL', 'VAR']]
        ]).groupby(['Name', 'PlayerId'], as_index=False).sum().sort_values(by='VAR', ascending=False)
        
        print("[✔] SGP Data Prepared!")
    
    def prepare_data(self, df, player_type, sb_included=False, auc_calc = None):
        df = df.reset_index()
        if (player_type=='Hitter'):
            cols_included = list(range(2,8))
            if (not sb_included):
                cols_included.remove(5)
            
            df['Total_SGP'] = df.iloc[:, cols_included].sum(axis=1)
            df['Total_SGP_wSB'] = df.iloc[:, 2:8].sum(axis=1)
            
            df = df.sort_values(by="Total_SGP", ascending=False).reset_index()
            
            auc_calc = auc_calc.rename(columns={'POS': 'ELIG'})
            df = df.merge(auc_calc[['PlayerId', 'ELIG']], on='PlayerId', how='left')
            
            # Apply function to create POS column
            df['POS'] = df['ELIG'].apply(self.determine_pos)

            for pos in ["1B", "3B", "2B", "SS", "C", "OF", "DH"]:
                df[f"{pos}_count"] = (df["POS"] == pos).cumsum()
            
            df["1B/3B_count"] = df["1B_count"] + df["3B_count"]
            df["2B/SS_count"] = df["2B_count"] + df["SS_count"]
            
            df["UTIL"] = self.assign_util(df)
            
            df.to_excel('temp_players.xlsx')
            
            print("[*] Computing replacement level (RL)...")
            df = self.compute_replacement_level(df)
            
        else:
            df['Total_SGP'] = df.iloc[:, 2:8].sum(axis=1)
            df['Starter'] = np.where(df['GS'] > 5, 1, 0)
            
            df = df.sort_values(by="Total_SGP", ascending=False).reset_index()

            print("[*] Computing replacement level for pitchers...")
            starter_rl = df[df["Starter"] == 1]["Total_SGP"].iloc[108]
            reliever_rl = df[df["Starter"] == 0]["Total_SGP"].iloc[36]
            
            df["RL"] = df.apply(lambda row: starter_rl if row["Starter"] == 1 else reliever_rl, axis=1)
            df["VAR"] = df["Total_SGP"] - df["RL"]
        
        print(f"[✔] {player_type} data preparation complete!")
        return df
    
    def determine_pos(self, elig):
        if pd.isna(elig):
            return "ERROR"
        for pos in ["C", "2B", "OF", "SS", "3B", "1B", "DH"]:
            if pos in elig:
                return pos
        return "ERROR"

    def assign_util(self, df):
        """Assign UTIL position based on rostered thresholds"""
        print("[*] Determining UTIL players...")
        util_list = []
        util_count = 0

        for _, row in df.iterrows():
            if (
                row["POS"] == "DH" or
                (row["POS"] == "OF" and row["OF_count"] >= 61) or
                (row["POS"] == "C" and row["C_count"] >= 13) or
                (row["POS"] in ["1B", "3B"] and row["1B_count"] + row["3B_count"] >= 37) or
                (row["POS"] in ["2B", "SS"] and row["2B_count"] + row["SS_count"] >= 37)
            ):
                util_count += 1
                util_list.append(util_count)
            else:
                util_list.append("")

        return util_list

    def compute_replacement_level(self, df):
        """Compute RL per position, ensuring required number of rostered players"""
        print("[*] Identifying rostered universe...")
        
        def get_rostered_universe(df):
            """Find the lowest-ranked players that fill required positions"""
            rostered = []
            needed = {"CI": 36, "MI": 36, "C": 12, "OF": 60, "UTIL": 12}
            counts = {k: 0 for k in needed}

            # Sort by Total_SGP (higher SGP players prioritized)
            df_sorted = df.sort_values(by="Total_SGP", ascending=False).reset_index(drop=True)

            for _, row in df_sorted.iterrows():
                assigned = False
                pos = row["POS"]  # Use POS for primary bucket assignment

                # Assign based on primary POS
                if pos in ["1B", "3B"] and counts["CI"] < needed["CI"]:
                    counts["CI"] += 1
                    rostered.append(row)
                    assigned = True
                elif pos in ["2B", "SS"] and counts["MI"] < needed["MI"]:
                    counts["MI"] += 1
                    rostered.append(row)
                    assigned = True
                elif pos == "C" and counts["C"] < needed["C"]:
                    counts["C"] += 1
                    rostered.append(row)
                    assigned = True
                elif pos == "OF" and counts["OF"] < needed["OF"]:
                    counts["OF"] += 1
                    rostered.append(row)
                    assigned = True

                # If not assigned to a primary position, consider for UTIL
                if not assigned and counts["UTIL"] < needed["UTIL"]:
                    counts["UTIL"] += 1
                    rostered.append(row)

                # Stop if we've reached the total roster size
                if sum(counts.values()) >= 156:
                    break

            # Convert to DataFrame and sort by Total_SGP before returning
            rostered = pd.DataFrame(rostered).sort_values(by="Total_SGP", ascending=False).reset_index(drop=True)

            return pd.DataFrame(rostered)

        rostered_df = get_rostered_universe(df)
        rostered_df.to_excel("rostered.xlsx")
        print(f"[✔] Rostered universe identified ({len(rostered_df)} players).")

        print("[*] Finding worst rostered player per position...")
        worst_rostered = {}
        for pos in ["1B", "3B", "2B", "SS", "C", "OF"]:
            worst_player = rostered_df[rostered_df["ELIG"].str.contains(pos, na=False)].nsmallest(1, "Total_SGP")
            if not worst_player.empty:
                worst_rostered[pos] = worst_player["Total_SGP"].values[0]
                print(f"[DEBUG] Worst rostered {pos}: {worst_player['Name'].values[0]} (SGP: {worst_rostered[pos]})")
            else:
                worst_rostered[pos] = float("inf")

        print("[*] Finding best non-rostered player per position...")
        best_replacements = {}
        for pos in worst_rostered.keys():
            best_player = df[~df["PlayerId"].isin(rostered_df["PlayerId"]) & df["ELIG"].str.contains(pos, na=False)].nlargest(1, "Total_SGP")
            if not best_player.empty:
                best_replacements[pos] = best_player["Total_SGP"].values[0]
                print(f"[DEBUG] Best non-rostered {pos}: {best_player['Name'].values[0]} (SGP: {best_replacements[pos]})")
            else:
                best_replacements[pos] = float("-inf")

        print("[*] Computing RL for each position...")
        rl_values = {
            pos: (best_replacements[pos] + worst_rostered[pos]) / 2
            if worst_rostered[pos] != float("inf") and best_replacements[pos] != float("-inf")
            else float("inf")
            for pos in worst_rostered
        }

        max_worst = max([val for val in worst_rostered.values() if val != float("inf")], default=float("inf"))
        max_best = max([val for val in best_replacements.values() if val != float("-inf")], default=float("-inf"))
    
        rl_values["DH"] = (max_best+max_worst)/2  # Take the maximum of both

        print(f"[DEBUG] DH RL penalized to: {rl_values['DH']} (Max of all worst and best non-rostered)")
        
        df["RL"] = df["ELIG"].apply(lambda elig: min(rl_values.get(pos, float("inf")) for pos in elig.split("/")))
        df["VAR"] = df["Total_SGP"] - df["RL"]
        
        print("[✔] Replacement level calculation complete!")
        return df
    
    def export_sgp(self,sb):
        print("[*] Exporting SGP Results...")
        
        SAVE_FOLDER = os.path.join(os.getcwd(), "results")
        os.makedirs(SAVE_FOLDER,exist_ok=True)
        
        sb_string = "_sb_included" if sb else ""
        file_name = f"results/SGP_Results_{self.suffix}{sb_string}.xlsx"
        
        with pd.ExcelWriter(file_name) as writer:
            self.hitters_df[['Name', 'PlayerId', 'PA', 'SGP_R', 'SGP_HR', 'SGP_RBI', 'SGP_SB', 'SGP_OBP', 'SGP_SLG', 'Total_SGP_wSB', 'Total_SGP', 'RL', 'VAR']].to_excel(writer, sheet_name='Hitters', index=False)
            self.pitchers_df[['Name', 'PlayerId', 'IP', 'SGP_SO', 'SGP_QS', 'SGP_SV_HLD', 'SGP_ERA', 'SGP_WHIP', 'SGP_K/BB', 'Total_SGP', 'RL', 'VAR']].to_excel(writer, sheet_name='Pitchers', index=False)
            self.combined_df.to_excel(writer, sheet_name='Combined', index=False)
        
        wb = load_workbook(file_name)
        ws_pitchers = wb["Pitchers"]  # Select the "Pitchers" sheet
        ws_combined = wb["Combined"]  # Select the "Combined" sheet

        # Define formatting (blue font & underlined)
        blue_underlined = Font(color="0000FF", underline="single", bold=True)

        # Get the list of pitchers from the "Combined" sheet (Name & PlayerId)
        combined_pitchers = set()
        for row in ws_pitchers.iter_rows(min_row=2, values_only=True):  # Skip header
            name, player_id, *_ = row  # Extract Name & PlayerId
            combined_pitchers.add((name, player_id))

        # Apply formatting only to pitchers who appear in the Combined sheet
        for row in ws_combined.iter_rows(min_row=2, max_row=ws_combined.max_row):
            name_cell, player_id_cell = row[0], row[1]  # Assuming Name is in column A, PlayerId in B
            if (name_cell.value, player_id_cell.value) in combined_pitchers:
                for cell in row:  # Apply formatting to the whole row
                    cell.font = blue_underlined

        wb.save(file_name)
        
        print(f"[✔] Exported SGP Results to {file_name}")
